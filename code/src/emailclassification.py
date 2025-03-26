#installing ollama
#!pip install ollama
!curl -fsSL https://ollama.com/install.sh | sh


#starting ollama server locally
import subprocess
import time
process = subprocess.Popen("ollama serve", shell=True)
time.sleep(5)  # Wait for 5 seconds


#pulling llama3 using ollama
!ollama pull llama3.1:8b


#testing llama 3 is availble
!ollama list


!pip install --upgrade langchain langchain_community faiss-cpu eml_parser pypdf transformers accelerate Flask pyngrok

import os
import json
import torch
from langchain.llms import Ollama
from langchain.embeddings import HuggingFaceEmbeddings
from langchain.vectorstores import FAISS
from langchain.document_loaders import PyPDFLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.chains import LLMChain
from langchain.prompts import PromptTemplate
import faiss
import json
import hashlib
import numpy as np
import email
import os
import eml_parser
from datetime import datetime
from langchain.embeddings import HuggingFaceEmbeddings
from langchain.llms import Ollama
from scipy.spatial.distance import cosine
from bs4 import BeautifulSoup
from langchain_community.vectorstores import FAISS
import configuration
import email
from bs4 import BeautifulSoup
import os
import tempfile
from langchain.document_loaders import PyPDFLoader
from openpyxl import load_workbook
import re
import json

# Load LLaMA 3 Model
llm = Ollama(model="llama3.1:8b")

# Load embedding model
embedding_model = HuggingFaceEmbeddings(model_name="sentence-transformers/all-mpnet-base-v2")

# Function to extract text from PDFs
def extract_text_from_pdf(pdf_paths):
    """Loads PDF documents using PyPDFLoader and returns a list of documents."""
    documents = []
    for pdf_path in pdf_paths:
        loader = PyPDFLoader(pdf_path)
        documents.extend(loader.load())
    text = ""
    for page in documents:
        text += page.page_content
    return text  # Load and extend the documents list

def extract_text_from_excel(excel_path):
    """Extracts text from an Excel file."""
    from openpyxl import load_workbook

    attachment_text = ""
    workbook = load_workbook(excel_path, read_only=True)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows():
            for cell in row:
                attachment_text += str(cell.value) + " "
    return attachment_text

# Function to extract text from EML files
def extract_email_data(eml_path):
    """
    Extracts data from an EML file, including subject, from, to, attachments,
    body text, and text from PDF and Excel attachments.
    """
    with open(eml_path, 'rb') as f:
        msg = email.message_from_binary_file(f)

    # Extract subject, from, and to
    subject = msg['Subject']
    sender = msg['From']
    recipients = msg['To']  # Can be multiple recipients

    # Extract attachments and their content
    attachments = []
    for part in msg.walk():
        if part.get_content_maintype() == 'multipart':
            continue
        if part.get('Content-Disposition') is None:
            continue

        filename = part.get_filename()
        if filename:
            # Save attachment to temporary file
            with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                temp_file.write(part.get_payload(decode=True))
                temp_file_path = temp_file.name

            # Extract text based on file type
            attachment_text = ""
            if filename.endswith(".pdf"):
                attachment_text += extract_text_from_pdf([temp_file_path])
            elif filename.endswith((".xlsx", ".xls")):
                attachment_text += extract_text_from_excel(temp_file_path)

            attachments.append({
                "filename": filename,
                "content": attachment_text  # Include attachment content
            })

            # Remove temporary file
            os.remove(temp_file_path)

    # Extract email body text
    email_text = ''
    for part in msg.walk():
        if part.get_content_type() == 'text/plain':
            email_text += part.get_payload(decode=True).decode('utf-8', errors='ignore')
        elif part.get_content_type() == 'text/html':
            html_content = part.get_payload(decode=True).decode('utf-8', errors='ignore')
            soup = BeautifulSoup(html_content, 'html.parser')
            email_text += soup.get_text()

    # Return data as a dictionary
    return {
        "subject": subject,
        "from": sender,
        "to": recipients,
        "attachments": attachments,
        "subject": subject,
        "body": email_text
    }


def extract_json_from_string(text):
    match = re.search(r'\{(.*?)\}', text, re.DOTALL)
    if match:
        json_string = match.group(0)
        try:
            return json.loads(json_string)
        except json.JSONDecodeError:
            return None  # Handle invalid JSON
    return None  # No match found
    

#Check for duplicates
def check_duplicate(email_text, vectorstore, threshold=0.8):

    # Create embedding for the current email
    email_embedding = embedding_model.embed_query(email_text)
    print("Current email embedding:", email_embedding) # Print for debugging

    # Search for similar emails in FAISS
    # Use similarity_search_with_score to get scores
    similar_emails_with_scores = vectorstore.similarity_search_with_score(email_text, k=5)
    print("Similar emails:", similar_emails_with_scores) # Print for debugging

    # Check for duplicates based on similarity score
    # Access the score from the tuple
    if similar_emails_with_scores and len(similar_emails_with_scores) > 1 and similar_emails_with_scores[1][1] > threshold:
        # Access metadata and score appropriately
        # Check if 'source' key exists before accessing it
        if 'source' in similar_emails_with_scores[1][0].metadata:
            print("Found duplicate with score:", similar_emails_with_scores[1][1]) # Print for debugging
            return True, similar_emails_with_scores[1][0].metadata['source'], similar_emails_with_scores[1][1]
        else:
            print("Found potential duplicate but missing 'source' metadata, score:", similar_emails_with_scores[1][1]) # Print for debugging
            return True, None, similar_emails_with_scores[1][1] # Handle case where 'source' key is missing and return score
    else:
        # If not a duplicate, return score which will be below threshold
        print("No duplicate found, score:", similar_emails_with_scores[1][1] if similar_emails_with_scores and len(similar_emails_with_scores) > 1 else 0.0) # Print for debugging
        return False, None, similar_emails_with_scores[1][1] if similar_emails_with_scores and len(similar_emails_with_scores) > 1 else 0.0


def store_email(email_text, vectorstore):
    """Stores an email in the FAISS vectorstore."""

    # Create embedding for the current email
    email_embedding = embedding_model.embed_query(email_text)

    # Add the email to the vectorstore
    # Instead of dictionary, use a Document object
    from langchain.schema import Document
    vectorstore.add_documents([Document(page_content=email_text, metadata={'embedding': email_embedding, 'source': email_text})])


# Function to extract key data points dynamically
def extract_key_data(text):
    prompt = f"""
    Extract key data points from the following text:

    Text: "{text}"

    Identify fields such as {configuration.keyMetrics}
    Respond in JSON format like this:
    {{
        "sender_name": "string",
        "amount": "string",
        "expiration_date": "string",
        "date_to_transfer": "string"
    }}
    """

    response = llm.invoke(prompt)
    try:
        extracted_data = extract_json_from_string(response)
    except json.JSONDecodeError:
        extracted_data = {}
    return extracted_data


def extract_requests_json(llm_response):
  """Extracts the 'requests' JSON from the LLM response using regex."""
  match = re.search(r'"requests":\s*(\[.*?\])', llm_response, re.DOTALL)
  if match:
    requests_json_str = match.group(1)
    try:
      requests_json = json.loads(requests_json_str)
      return requests_json
    except json.JSONDecodeError:
      print("Error: Invalid JSON format for 'requests'")
      return []
  else:
    print("Error: 'requests' key not found in LLM response")
    return []


# Function to extract key data points dynamically
def classify_email(text):

    prompt = f"""
You are a commercial loan servicing assistant. Your task is to analyze the provided email body and identify all service requests mentioned, classifying each into a specific category and subtype.

Email Body:
{text}

Available Request Types and Subtypes:
{configuration.requestTypesAndSubTypes}

Instructions:
1. Thoroughly review the email body and identify all instances where a service request is being made.
2. For each request, determine the most appropriate request category from the provided list.
3. If applicable, select the relevant subrequest type based on the specifics of the request.
4. Base your confidence on factors like:
    - Presence of keywords related to the request category/subtype.
    - Overall semantic similarity between the request and the category/subtype description.
    - The strength and clarity of the request expressed in the email.
5. If a single sentence or phrase implies multiple requests, prioritize the most specific and explicit request.
6. If there are overlapping or ambiguous requests, use your best judgment to classify them based on the overall context of the email.

Return your classification in the following JSON format:

{{
  "requests": [
    {{
      "request_category": "Selected Category",
      "subrequest_type": "Selected Subtype",
      "confidence_reason": "Brief explanation for your confidence level"
      "confidence score percentage": ""
    }},
    {{
      "request_category": "Selected Category",
      "subrequest_type": "Selected Subtype",
      "confidence_reason": "Brief explanation for your confidence level"
      "confidence score percentage": ""
    }},
    // ... more requests
  ]
}}
"""

    response = llm.invoke(prompt)
    #print("llm res", response)
    try:
        extracted_data = extract_requests_json(response)
    except json.JSONDecodeError:
        extracted_data = []
    return extracted_data

#initializing vector store
example_texts = ["This is an example text for initialization."]
vectorstore = FAISS.from_texts(example_texts, embedding_model)

all_docs = vectorstore.similarity_search("", k=vectorstore.index.ntotal)
print(len(all_docs))

def process_eml_file(eml_file_path):
    """Processes an EML file and returns the results."""
    results = []

    if eml_file_path.endswith(".eml"):
        email_text = extract_email_data(eml_file_path)
        email_content = email_text['subject'] + email_text['body']
        email_attachments_content = email_text['attachments']

        is_dup, source, score = check_duplicate(email_content, vectorstore)

        response_json = {}
        if is_dup:
            response_json = {
                "email_text": email_content[:100] + "...",
                "classification": "duplicate",
                "duplicate": is_dup,
            }
        else:
            classification = classify_email(email_content)
            extracted_data = extract_key_data(email_content)
            store_email(email_content, vectorstore)
            response_json = {
                "email_text": email_content[:100] + "...",
                "classification": classification,
                "duplicate": is_dup,
                "key_data": extracted_data
            }
        results.append(response_json)

        print("resuls", json.dumps(results, indent=2))

    return results

process_eml_file("email.eml")

from flask import Flask, request, jsonify
import json
from pyngrok import ngrok

app = Flask(__name__)

@app.route('/process_email', methods=['POST'])
def process_email():
    print("into process mail")
    if 'eml_file' not in request.files:
        return jsonify({'error': 'No EML file provided'}), 400

    print("eml file exists")
    eml_file = request.files['eml_file']
    eml_file.save('temp_email.eml')  # Save temporarily
    print("saved to temp file")
    results = process_eml_file('temp_email.eml')  # Call the processing method
    print("results received ", results)
    response_data = {
        "message": "Email processed successfully",
        "results": results
    }

    return jsonify(response_data), 200

if __name__ == '__main__':
     ngrok.set_auth_token("2t6u6zBAZFZh2JGODqQ0s0MyQt8_7XXGdkuk8Mm3UhCdw73WA") # signup and get auth token from https://dashboard.ngrok.com/get-started/your-authtoken
     public_url = ngrok.connect(5000).public_url
     print(f" * Running on {public_url} (Press CTRL+C to quit)")
     app.run(host='0.0.0.0', port=5000, debug=True)