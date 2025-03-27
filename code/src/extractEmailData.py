import email
from bs4 import BeautifulSoup
import os
import tempfile
from langchain.document_loaders import PyPDFLoader
from openpyxl import load_workbook

# Function to extract text from PDFs
def extract_text_from_pdf(pdf_paths):
    """Loads PDF documents using PyPDFLoader and returns a list of documents."""
    documents = []
    for pdf_path in pdf_paths:
        loader = PyPDFLoader(pdf_path)
        documents.extend(loader.load())
    text = ""
    for page in documents:
        text += page.page_content
    return text  # Load and extend the documents list

def extract_text_from_excel(excel_path):
    """Extracts text from an Excel file."""
    from openpyxl import load_workbook

    attachment_text = ""
    workbook = load_workbook(excel_path, read_only=True)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows():
            for cell in row:
                attachment_text += str(cell.value) + " "
    return attachment_text

# Function to extract text from EML files
def extract_email_data(eml_path):
    """
    Extracts data from an EML file, including subject, from, to, attachments,
    body text, and text from PDF and Excel attachments.
    """
    with open(eml_path, 'rb') as f:
        msg = email.message_from_binary_file(f)

    # Extract subject, from, and to
    subject = msg['Subject']
    sender = msg['From']
    recipients = msg['To']  # Can be multiple recipients

    # Extract attachments and their content
    attachments = []
    for part in msg.walk():
        if part.get_content_maintype() == 'multipart':
            continue
        if part.get('Content-Disposition') is None:
            continue

        filename = part.get_filename()
        if filename:
            # Save attachment to temporary file
            with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                temp_file.write(part.get_payload(decode=True))
                temp_file_path = temp_file.name

            # Extract text based on file type
            attachment_text = ""
            if filename.endswith(".pdf"):
                attachment_text += extract_text_from_pdf([temp_file_path])
            elif filename.endswith((".xlsx", ".xls")):
                attachment_text += extract_text_from_excel(temp_file_path)

            attachments.append({
                "filename": filename,
                "content": attachment_text  # Include attachment content
            })

            # Remove temporary file
            os.remove(temp_file_path)

    # Extract email body text
    email_text = ''
    for part in msg.walk():
        if part.get_content_type() == 'text/plain':
            email_text += part.get_payload(decode=True).decode('utf-8', errors='ignore')
        elif part.get_content_type() == 'text/html':
            html_content = part.get_payload(decode=True).decode('utf-8', errors='ignore')
            soup = BeautifulSoup(html_content, 'html.parser')
            email_text += soup.get_text()

    # Return data as a dictionary
    return {
        "subject": subject,
        "from": sender,
        "to": recipients,
        "attachments": attachments,
        "subject": subject,
        "body": email_text
    }